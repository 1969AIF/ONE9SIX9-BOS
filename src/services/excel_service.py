"""Excel workbook management for ONE9SIX9 BOS."""

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

WORKBOOK_NAME = "ONE9SIX9.xlsx"

WORKSHEET_HEADERS: dict[str, list[str]] = {
    "Dashboard": ["Metric", "Value", "Period", "Last Updated"],
    "Expenses": [
        "ID",
        "Date",
        "Supplier",
        "Project",
        "Category",
        "Invoice/Receipt Number",
        "Description",
        "Amount Excl. VAT",
        "VAT",
        "Total Incl. VAT",
        "Status",
        "Notes",
    ],
    "Income": [
        "ID",
        "Date",
        "Description",
        "Project",
        "Client",
        "Amount",
        "Status",
        "Notes",
    ],
    "Projects": [
        "Project Code",
        "Project Name",
        "Status",
        "Client",
        "Start Date",
        "End Date",
        "Budget",
        "Notes",
    ],
    "Suppliers": [
        "Supplier Name",
        "VAT Number",
        "Contact Person",
        "Telephone",
        "Email",
        "Address",
        "Category",
        "Notes",
    ],
    "Categories": ["Category Code", "Category Name", "Type", "Description"],
    "Settings": ["Key", "Value", "Description"],
}

PROJECT_CODE_HEADERS = ("Project Code", "ID")
PROJECT_NAME_HEADERS = ("Project Name", "Name")
PROJECT_STATUS_HEADER = "Status"

SUPPLIER_NAME_HEADERS = ("Supplier Name", "Name")
SUPPLIER_VAT_HEADER = "VAT Number"
SUPPLIER_CONTACT_HEADERS = ("Contact Person",)
SUPPLIER_TELEPHONE_HEADERS = ("Telephone", "Phone")
SUPPLIER_EMAIL_HEADERS = ("Email",)

CATEGORY_CODE_HEADERS = ("Category Code", "Code", "ID")
CATEGORY_NAME_HEADERS = ("Category Name", "Name")

EXPENSE_ID_HEADERS = ("ID",)
EXPENSE_DATE_HEADERS = ("Date",)
EXPENSE_SUPPLIER_HEADERS = ("Supplier", "Supplier Name")
EXPENSE_PROJECT_HEADERS = ("Project",)
EXPENSE_CATEGORY_HEADERS = ("Category",)
EXPENSE_INVOICE_HEADERS = (
    "Invoice/Receipt Number",
    "Invoice Number",
    "Receipt Number",
    "Invoice",
    "Receipt",
)
EXPENSE_DESCRIPTION_HEADERS = ("Description",)
EXPENSE_AMOUNT_HEADERS = ("Amount Excl. VAT", "Amount", "Amount Excluding VAT")
EXPENSE_VAT_HEADERS = ("VAT",)
EXPENSE_TOTAL_HEADERS = ("Total Incl. VAT", "Total", "Total Including VAT")


@dataclass
class Project:
    """A project record from the Projects worksheet."""

    project_code: str
    project_name: str
    status: str


@dataclass
class Supplier:
    """A supplier record from the Suppliers worksheet."""

    supplier_name: str
    vat_number: str
    contact_person: str
    telephone: str
    email: str


@dataclass
class Category:
    """A category record from the Categories worksheet."""

    category_code: str
    category_name: str


@dataclass
class Expense:
    """An expense record from the Expenses worksheet."""

    date: str
    supplier: str
    project: str
    category: str
    invoice_number: str
    description: str
    amount_ex_vat: str
    vat: str
    total: str


class ExcelService:
    """Creates, opens, and manages the ONE9SIX9 Excel workbook."""

    _workbook: Workbook | None = None
    _workbook_path: Path | None = None

    @classmethod
    def initialize(cls, project_root: Path) -> None:
        """Ensure the workbook exists and is loaded."""
        cls._workbook_path = project_root / WORKBOOK_NAME

        if cls._workbook_path.exists():
            cls._workbook = load_workbook(cls._workbook_path)
        else:
            cls._workbook = cls._create_workbook()
            cls._workbook.save(cls._workbook_path)

    @classmethod
    def get_workbook(cls) -> Workbook:
        """Return the loaded workbook."""
        if cls._workbook is None:
            raise RuntimeError("ExcelService has not been initialized.")
        return cls._workbook

    @classmethod
    def get_workbook_path(cls) -> Path:
        """Return the path to the workbook file."""
        if cls._workbook_path is None:
            raise RuntimeError("ExcelService has not been initialized.")
        return cls._workbook_path

    @classmethod
    def save(cls) -> None:
        """Persist the current workbook to disk."""
        if cls._workbook is None or cls._workbook_path is None:
            raise RuntimeError("ExcelService has not been initialized.")
        cls._workbook.save(cls._workbook_path)

    @classmethod
    def get_projects(cls) -> list[Project]:
        """Return all projects from the Projects worksheet."""
        sheet = cls._get_projects_sheet()
        column_map = cls._get_project_column_map(sheet)

        projects: list[Project] = []
        for row in range(2, sheet.max_row + 1):
            project_code = cls._cell_value(sheet, row, column_map["code"])
            project_name = cls._cell_value(sheet, row, column_map["name"])
            status = cls._cell_value(sheet, row, column_map["status"])

            if not project_code and not project_name:
                continue

            projects.append(
                Project(
                    project_code=project_code,
                    project_name=project_name,
                    status=status or "Active",
                )
            )

        return projects

    @classmethod
    def add_project(
        cls,
        project_code: str,
        project_name: str,
        status: str,
    ) -> None:
        """Append a new project row and save the workbook."""
        sheet = cls._get_projects_sheet()
        column_map = cls._get_project_column_map(sheet)
        next_row = sheet.max_row + 1

        sheet.cell(row=next_row, column=column_map["code"], value=project_code)
        sheet.cell(row=next_row, column=column_map["name"], value=project_name)
        sheet.cell(row=next_row, column=column_map["status"], value=status)

        cls.save()

    @classmethod
    def get_suppliers(cls) -> list[Supplier]:
        """Return all suppliers from the Suppliers worksheet."""
        sheet = cls._get_suppliers_sheet()
        column_map = cls._get_supplier_column_map(sheet)

        suppliers: list[Supplier] = []
        for row in range(2, sheet.max_row + 1):
            supplier_name = cls._cell_value(sheet, row, column_map["name"])

            if not supplier_name:
                continue

            suppliers.append(
                Supplier(
                    supplier_name=supplier_name,
                    vat_number=cls._optional_cell_value(
                        sheet, row, column_map["vat_number"]
                    ),
                    contact_person=cls._cell_value(
                        sheet, row, column_map["contact_person"]
                    ),
                    telephone=cls._cell_value(sheet, row, column_map["telephone"]),
                    email=cls._cell_value(sheet, row, column_map["email"]),
                )
            )

        return suppliers

    @classmethod
    def add_supplier(
        cls,
        supplier_name: str,
        vat_number: str,
        contact_person: str,
        telephone: str,
        email: str,
    ) -> None:
        """Append a new supplier row and save the workbook."""
        sheet = cls._get_suppliers_sheet()
        column_map = cls._get_supplier_column_map(sheet, ensure_vat_column=True)
        next_row = sheet.max_row + 1

        sheet.cell(row=next_row, column=column_map["name"], value=supplier_name)
        if column_map["vat_number"] > 0:
            sheet.cell(row=next_row, column=column_map["vat_number"], value=vat_number)
        sheet.cell(
            row=next_row, column=column_map["contact_person"], value=contact_person
        )
        sheet.cell(row=next_row, column=column_map["telephone"], value=telephone)
        sheet.cell(row=next_row, column=column_map["email"], value=email)

        cls.save()

    @classmethod
    def get_categories(cls) -> list[Category]:
        """Return all categories from the Categories worksheet."""
        sheet = cls._get_categories_sheet()
        column_map = cls._get_category_column_map(sheet)

        categories: list[Category] = []
        for row in range(2, sheet.max_row + 1):
            category_code = cls._cell_value(sheet, row, column_map["code"])
            category_name = cls._cell_value(sheet, row, column_map["name"])

            if not category_code and not category_name:
                continue

            categories.append(
                Category(
                    category_code=category_code,
                    category_name=category_name,
                )
            )

        return categories

    @classmethod
    def add_category(
        cls,
        category_code: str,
        category_name: str,
    ) -> None:
        """Append a new category row and save the workbook."""
        sheet = cls._get_categories_sheet()
        column_map = cls._get_category_column_map(sheet)
        next_row = sheet.max_row + 1

        sheet.cell(row=next_row, column=column_map["code"], value=category_code)
        sheet.cell(row=next_row, column=column_map["name"], value=category_name)

        cls.save()

    @classmethod
    def _get_categories_sheet(cls) -> Worksheet:
        return cls.get_workbook()["Categories"]

    @classmethod
    def _get_category_column_map(cls, sheet: Worksheet) -> dict[str, int]:
        headers = [str(cell.value) if cell.value is not None else "" for cell in sheet[1]]

        return {
            "code": cls._find_column(headers, CATEGORY_CODE_HEADERS),
            "name": cls._find_column(headers, CATEGORY_NAME_HEADERS),
        }

    @classmethod
    def get_expenses(cls) -> list[Expense]:
        """Return all expenses from the Expenses worksheet."""
        sheet = cls._get_expenses_sheet()
        column_map = cls._get_expense_column_map(sheet)

        expenses: list[Expense] = []
        for row in range(2, sheet.max_row + 1):
            date = cls._cell_value(sheet, row, column_map["date"])
            supplier = cls._cell_value(sheet, row, column_map["supplier"])
            project = cls._cell_value(sheet, row, column_map["project"])
            category = cls._cell_value(sheet, row, column_map["category"])
            invoice_number = cls._cell_value(sheet, row, column_map["invoice"])
            description = cls._cell_value(sheet, row, column_map["description"])
            amount = cls._cell_value(sheet, row, column_map["amount"])
            vat = cls._cell_value(sheet, row, column_map["vat"])
            total = cls._cell_value(sheet, row, column_map["total"])

            if not date and not supplier and not project and not description:
                continue

            expenses.append(
                Expense(
                    date=date,
                    supplier=supplier,
                    project=project,
                    category=category,
                    invoice_number=invoice_number,
                    description=description,
                    amount_ex_vat=amount,
                    vat=vat,
                    total=total,
                )
            )

        return expenses

    @classmethod
    def add_expense(
        cls,
        *,
        date: str,
        supplier: str,
        project: str,
        category: str,
        invoice_number: str,
        description: str,
        amount_ex_vat: float,
        vat: float,
        total: float,
    ) -> None:
        """Append a new expense row and save the workbook."""
        sheet = cls._get_expenses_sheet()
        column_map = cls._get_expense_column_map(sheet)
        next_row = sheet.max_row + 1

        sheet.cell(row=next_row, column=column_map["id"], value=cls._next_expense_id(sheet, column_map))
        sheet.cell(row=next_row, column=column_map["date"], value=date)
        sheet.cell(row=next_row, column=column_map["supplier"], value=supplier)
        sheet.cell(row=next_row, column=column_map["project"], value=project)
        sheet.cell(row=next_row, column=column_map["category"], value=category)
        sheet.cell(row=next_row, column=column_map["invoice"], value=invoice_number)
        sheet.cell(row=next_row, column=column_map["description"], value=description)
        sheet.cell(row=next_row, column=column_map["amount"], value=amount_ex_vat)
        sheet.cell(row=next_row, column=column_map["vat"], value=vat)
        sheet.cell(row=next_row, column=column_map["total"], value=total)

        cls.save()

    @classmethod
    def _get_expenses_sheet(cls) -> Worksheet:
        return cls.get_workbook()["Expenses"]

    @classmethod
    def _get_expense_column_map(cls, sheet: Worksheet) -> dict[str, int]:
        headers = [str(cell.value) if cell.value is not None else "" for cell in sheet[1]]

        return {
            "id": cls._find_column(headers, EXPENSE_ID_HEADERS),
            "date": cls._find_column(headers, EXPENSE_DATE_HEADERS),
            "supplier": cls._find_column(headers, EXPENSE_SUPPLIER_HEADERS),
            "project": cls._find_column(headers, EXPENSE_PROJECT_HEADERS),
            "category": cls._find_column(headers, EXPENSE_CATEGORY_HEADERS),
            "invoice": cls._find_column(headers, EXPENSE_INVOICE_HEADERS),
            "description": cls._find_column(headers, EXPENSE_DESCRIPTION_HEADERS),
            "amount": cls._find_column(headers, EXPENSE_AMOUNT_HEADERS),
            "vat": cls._find_column(headers, EXPENSE_VAT_HEADERS),
            "total": cls._find_column(headers, EXPENSE_TOTAL_HEADERS),
        }

    @classmethod
    def _next_expense_id(cls, sheet: Worksheet, column_map: dict[str, int]) -> int:
        max_id = 0
        for row in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row, column=column_map["id"]).value
            if isinstance(value, (int, float)):
                max_id = max(max_id, int(value))
        return max_id + 1

    @classmethod
    def _get_projects_sheet(cls) -> Worksheet:
        return cls.get_workbook()["Projects"]

    @classmethod
    def _get_project_column_map(cls, sheet: Worksheet) -> dict[str, int]:
        headers = [str(cell.value) if cell.value is not None else "" for cell in sheet[1]]

        return {
            "code": cls._find_column(headers, PROJECT_CODE_HEADERS),
            "name": cls._find_column(headers, PROJECT_NAME_HEADERS),
            "status": cls._find_column(headers, (PROJECT_STATUS_HEADER,)),
        }

    @classmethod
    def _get_suppliers_sheet(cls) -> Worksheet:
        return cls.get_workbook()["Suppliers"]

    @classmethod
    def _get_supplier_column_map(
        cls,
        sheet: Worksheet,
        *,
        ensure_vat_column: bool = False,
    ) -> dict[str, int]:
        headers = [str(cell.value) if cell.value is not None else "" for cell in sheet[1]]

        column_map = {
            "name": cls._find_column(headers, SUPPLIER_NAME_HEADERS),
            "contact_person": cls._find_column(headers, SUPPLIER_CONTACT_HEADERS),
            "telephone": cls._find_column(headers, SUPPLIER_TELEPHONE_HEADERS),
            "email": cls._find_column(headers, SUPPLIER_EMAIL_HEADERS),
        }

        if SUPPLIER_VAT_HEADER in headers:
            column_map["vat_number"] = headers.index(SUPPLIER_VAT_HEADER) + 1
        elif ensure_vat_column:
            column_map["vat_number"] = len(headers) + 1
            sheet.cell(row=1, column=column_map["vat_number"], value=SUPPLIER_VAT_HEADER)
        else:
            column_map["vat_number"] = 0

        return column_map

    @classmethod
    def _optional_cell_value(cls, sheet: Worksheet, row: int, column: int) -> str:
        if column <= 0:
            return ""
        return cls._cell_value(sheet, row, column)

    @staticmethod
    def _find_column(headers: list[str], candidates: tuple[str, ...]) -> int:
        for candidate in candidates:
            if candidate in headers:
                return headers.index(candidate) + 1
        raise ValueError(f"Required column not found: {candidates}")

    @staticmethod
    def _cell_value(sheet: Worksheet, row: int, column: int) -> str:
        value = sheet.cell(row=row, column=column).value
        return str(value).strip() if value is not None else ""

    @classmethod
    def _create_workbook(cls) -> Workbook:
        """Build a new workbook with all required worksheets and headers."""
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        for sheet_name, headers in WORKSHEET_HEADERS.items():
            worksheet = workbook.create_sheet(title=sheet_name)
            cls._write_headers(worksheet, headers)

        return workbook

    @staticmethod
    def _write_headers(worksheet: Worksheet, headers: list[str]) -> None:
        """Write column headers to the first row of a worksheet."""
        for column_index, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=column_index, value=header)
